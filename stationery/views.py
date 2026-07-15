from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse
from django.db.models import F
from openpyxl import Workbook

from .forms import RequisitionForm
from .models import Requisition, Inventory


def custom_login(request):
    if request.user.is_authenticated:
        return redirect("create_requisition")

    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)

        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]

            user = authenticate(
                request,
                username=username,
                password=password,
            )

            if user is not None:
                login(request, user)

                if user.is_staff:
    return redirect("/admin/")

                return redirect("create_requisition")

        messages.error(request, "Invalid username or password.")

    else:
        form = AuthenticationForm()

    return render(
        request,
        "registration/login.html",
        {"form": form},
    )


@login_required
def create_requisition(request):

    if request.method == "POST":

        form = RequisitionForm(request.POST)

        if form.is_valid():

            requisition = form.save(commit=False)
            requisition.requested_by = request.user.username
            requisition.save()

            messages.success(
                request,
                "Request submitted successfully."
            )

            return redirect("create_requisition")

        else:

            for field in form:
                for error in field.errors:
                    messages.error(
                        request,
                        f"{field.label}: {error}"
                    )

            for error in form.non_field_errors():
                messages.error(request, error)

    else:
        form = RequisitionForm()

    return render(
        request,
        "requisition_form.html",
        {
            "form": form
        }
    )


@login_required
def view_requests(request):

    requests = Requisition.objects.filter(
        requested_by=request.user.username
    ).order_by("-created_at")

    return render(
        request,
        "request_status.html",
        {
            "requests": requests
        }
    )


@login_required
def dashboard(request):

    total_requests = Requisition.objects.count()

    pending_requests = Requisition.objects.filter(
        status="Pending"
    ).count()

    approved_requests = Requisition.objects.filter(
        status="Approved"
    ).count()

    rejected_requests = Requisition.objects.filter(
        status="Rejected"
    ).count()

    low_stock_items = Inventory.objects.filter(
        available_qty__lte=F("minimum_qty")
    )

    low_stock_count = low_stock_items.count()

    context = {
        "total_requests": total_requests,
        "pending_requests": pending_requests,
        "approved_requests": approved_requests,
        "rejected_requests": rejected_requests,
        "low_stock_items": low_stock_items,
        "low_stock_count": low_stock_count,
    }

    return render(
        request,
        "dashboard.html",
        context,
    )


@login_required
def export_excel(request):

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Stationery Requests"

    headers = [
        "Department",
        "Item",
        "Requested By",
        "Requested Qty",
        "Approved Qty",
        "Status",
        "Reason",
        "Created At",
    ]

    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col).value = header

    row = 2

    for req in Requisition.objects.all().order_by("-created_at"):

        worksheet.cell(row=row, column=1).value = req.department
        worksheet.cell(row=row, column=2).value = req.item.item_name
        worksheet.cell(row=row, column=3).value = req.requested_by
        worksheet.cell(row=row, column=4).value = req.requested_qty
        worksheet.cell(row=row, column=5).value = req.approved_qty
        worksheet.cell(row=row, column=6).value = req.status
        worksheet.cell(row=row, column=7).value = req.reason
        worksheet.cell(row=row, column=8).value = req.created_at.strftime("%d-%m-%Y %H:%M")

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Stationery_Requests.xlsx"'
    )

    workbook.save(response)

    return response